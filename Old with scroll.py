from docxtpl import DocxTemplate
import openpyxl
from tkinter import *
from tkinter import messagebox
import os
from pathlib import Path
from tkinter import filedialog
from tkinter import ttk
import re
import faker
from random import randrange

path_sample = ''
path_table = ''
user_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents', 'Шаблоны')


def error(title, text):
    messagebox.showerror(title=title, message=text)


def info(title, text):
    messagebox.showinfo(title=title, message=text)


def open_sample_doc():
    global path_sample
    path_sample = filedialog.askopenfilename(title="Выбор шаблона для заполнения", initialdir=user_path,
                                             filetypes=(("Документы (*.docx)",
                                                         "*.docx"),
                                                        ("Все файлы", "*.*")))
    if path_sample != '':
        label3['foreground'] = 'green'


def open_table():
    global path_table
    path_table = filedialog.askopenfilename(title="Выбор таблицы для заполнения", initialdir=user_path,
                                            filetypes=(("Таблицы (*.xlsx)",
                                                        "*.xlsx"),
                                                       ("Все файлы", "*.*")))
    if path_table != '':
        label4['foreground'] = 'green'
        wb = openpyxl.load_workbook(path_table)
        sheet = wb.get_sheet_by_name('Доверенность')
        e_end.insert(0, sheet['M'][-1].coordinate)
        wb.close()


def get_dict():
    start = e_start.get().upper()
    end = e_end.get().upper()
    test = []
    wb = openpyxl.load_workbook(f'{path_table}')
    sheet = wb.active

    rows = 0

    try:
        for row in sheet[f'{start}':f'{end}']:
            rows += 1
            for cellObj in row:
                if cellObj.value == None or cellObj.value == '':
                    continue
                # if len(cellObj.value) > 60:
                #     if cellObj.value[60] == ' ':
                #         test.append(cellObj.value[:60])
                #         test.append(cellObj.value[60:])
                #     else:
                #         if cellObj.value[60] == ' ':
                test.append(cellObj.value)
                # print(f'Длина для {cellObj.value}: {len(cellObj.value)}')

        info('Обработка входных данных', "Данные из таблицы успешно загружены!\nПереходим к "
                                         "обработке шаблона.\n"
                                         f"Было обработано строк: {rows}")
        wb.close()
        return test
    except ValueError:
        error("Обработка входных данных", "Введите корректные коордиаты!")
        wb.close()
    except:
        error("Обработка входных данных", "Ошибка обработки данных!")
        wb.close()


def start(test):
    x = 0
    files = 0
    path = Path(user_path + '/Доверенности')

    if not path.exists():
        os.mkdir(path)

    try:
        while x < len(test):
            doc = DocxTemplate(f'{path_sample}')
            context = {'director': test[x + 2], 'doc': test[x + 3], 'job': test[x + 4], 'name_of_subject': test[x + 5],
                       'seria': test[x + 6], 'num': test[x + 7], 'date': test[x + 8], 'passport_from_1': test[x + 9],
                       'job_of_director': test[x + 10], 'name_of_director': test[x + 11]}
            doc.render(context)
            path_dirs = Path(f'{path}/{test[x]}')
            if not path_dirs.exists():
                os.mkdir(path_dirs)
            doc.save(f'{path}/{test[x]}/Доверенность ' + test[x + 1] + '.docx')
            x += 12  # прыжок на следующую строку
            files += 1
        messagebox.showinfo(title="Обработка входных данных", message="Шаблон успешно обработан.\n"
                                                                      f"Создано файлов: {files}.")
    except:
        error("Обработка входных данных", "Ошибка заполнения шаблона!")


def check_text():
    if e_start.get().isalnum() and e_end.get().isalnum():
        return True
    else:
        return False


def fill_files():
    if check_text():
        if path_sample != '' and path_table != '':
            start(get_dict())
            e_start.delete(0, END)
            e_end.delete(0, END)
        else:
            error("Ошибка входных данных", "Выберите шаблон и таблицу с данными!")
    else:
        error("Ошибка входных данных", "Введите начальную и конечную точку таблицы!")


def open_dir():
    path = Path(user_path + '/Доверенности')

    if path.exists():
        os.startfile(path)
    else:
        error("Открытие папки", "Каталог не существует!")


def open_custom_fill():
    path_xlsx = ''

    def check_entry():
        if gender_data != '' and e_fio.get() and e_director.get() and e_doc.get() and \
                e_job.get() and e_name.get() and e_seria.get() and e_number.get() and e_date.get() and \
                e_point_of_get.get() and e_directors_job.get() and e_directors_name.get() and \
                e_code.get() and e_birth.get() and e_place_of_birth.get() and gender.get() and e_full_fio.get() and \
                e_inn.get() and e_snils.get() and e_mail.get():
            return True
        else:
            return False

    def open_xlsx():
        nonlocal path_xlsx
        path_xlsx = filedialog.askopenfilename(title="Выберите вашу таблицу с данными", initialdir=user_path,
                                               filetypes=(("Таблицы (*.xlsx)",
                                                           "*.xlsx"),
                                                          ("Все файлы", "*.*")))
        if path_xlsx != '':
            label13['foreground'] = 'green'

    def open_xlsx_sample():
        path = Path(user_path)

        if path.exists():
            os.startfile(path)
        else:
            error("Открытие папки", "Каталог не существует!")

    def create_xlsx():
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet_1 = book.create_sheet("Доверенность")
        sheet_2 = book.create_sheet("Заявление")

        sheet_1.insert_rows(0)
        col_names = ['№', 'Город', 'На кого', 'Руководитель', 'Основание', 'Должность', 'ФИО', 'Серия паспорта',
                     'Номер паспорта', 'Дата выдачи', 'Кем выдан', 'Должность руководителя', 'ФИО руководителя']
        l = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
             'V', 'W', 'X', 'Y', 'Z']
        n = 0
        for i in col_names:
            sheet_1[f'{l[n]}1'].value = f'{col_names[n]}'
            sheet_1.column_dimensions[l[n]].width = 20
            n += 1

        sheet_2.insert_rows(0)
        col2_names = ['№', 'Серия', 'Номер', 'Дата выдачи', 'Код', 'Дата рождения', 'Место рождения', 'Пол',
                      'Фамилия', 'Имя', 'Отчество', 'ИНН', 'СНИЛС', 'Должность', 'Почта', 'Населенный пункт']
        k = 0
        for i in col2_names:
            sheet_2[f'{l[k]}1'].value = f'{col2_names[k]}'
            sheet_2.column_dimensions[l[k]].width = 20
            k += 1

        path = Path(user_path)
        if not path.exists():
            os.mkdir(path)

        sheet_1.auto_filter.ref = "B1:M999"
        sheet_2.auto_filter.ref = "B1:P999"

        book.save(f"{path}/Переменные.xlsx")
        info("Создание таблицы", 'Таблица успешно создана!')

    def clear_xlsx():
        if path_xlsx != '':
            question = messagebox.askokcancel(title="Очистка таблицы", message="Вы уверены, что хотите очистить "
                                                                               "таблицу?\nВсе "
                                                                               "данные в таблице будут безвозвратно удалены!")
            if question:
                book = openpyxl.load_workbook(path_xlsx)
                sheet = book["Доверенность"]
                sheet2 = book["Заявление"]
                last_row = len(list(sheet.rows))
                first_row = 1
                sheet.delete_rows(first_row, last_row)
                last_row2 = len(list(sheet2.rows))
                first_row2 = 1
                sheet2.delete_rows(first_row2, last_row2)
                del1 = last_row - first_row
                del2 = last_row2 - first_row2
                if del1 < 0:
                    del1 = 0
                if del2 < 0:
                    del2 = 0
                book.save(path_xlsx)
                info("Очистка таблицы", f'Успешно удалено строк: {del1} в листе "Доверенность" и '
                                        f'{del2} в листе "Заявление"')


        else:
            error("Ошибка входных данных", "Выберите таблицу с данными!")

    gender_data = ''

    def insert_data():
        clear_data()
        # e_city.insert(0, 'Красноярск')
        e_director.insert(0, 'Главного врача Филиной Натальи Григорьевны')
        e_doc.insert(0, 'Устава, утвержденного Приказом от 28.12.2021 №2586-орг')
        e_directors_job.insert(0, 'Главный врач')
        e_directors_name.insert(0, 'Н.Г. Филина')
        e_mail.insert(0, 'ikdomashenko@kkck.ru')

    def get_data():
        data = [[gender_data, e_fio.get(), e_director.get(), e_doc.get(), e_job.get(), e_name.get(), e_seria.get(),
                 e_number.get(), e_date.get(), e_point_of_get.get(), e_directors_job.get(), e_directors_name.get()]]
        return data

    def get_data2():
        temp = e_full_fio.get()
        temp2 = temp.split()
        if len(temp2) < 3:
            error("Ошибка входных данных", "Укажите верное ФИО (полное)!")
        else:
            fam = temp2[0]
            name = temp2[1]
            otch = temp2[2]
            temp = e_snils.get()
            snils = "".join(c for c in temp if c.isalnum())
            city = combo_city.get().rstrip()
            data2 = [
                [
                    e_seria.get(), e_number.get(), e_date.get(), e_code.get(), e_birth.get(), e_place_of_birth.get(),
                    gender.get(), fam, name, otch, e_inn.get(), snils, e_job.get(), e_mail.get(), city
                ]
            ]
            return data2

    def confirm_data():
        if path_xlsx == '':
            error("Ошибка входных данных", "Укажите таблицу данных!")
        else:
            if check_entry():
                wb = openpyxl.load_workbook(path_xlsx)
                sheet = wb["Доверенность"]
                temp = get_data()
                try:
                    number = int(sheet["A"][-1].value) + 1
                except:
                    number = 1
                data = temp
                data[0].insert(0, number)
                for row in data:
                    sheet.append(row)
                # wb1.save(path_xlsx)

                # wb2 = openpyxl.load_workbook(path_xlsx)
                sheet2 = wb["Заявление"]
                temp2 = get_data2()
                try:
                    number = int(sheet2["A"][-1].value) + 1
                except:
                    number = 1
                data2 = temp2
                data2[0].insert(0, number)
                for row in data2:
                    sheet2.append(row)
                wb.save(path_xlsx)
                info('Обработка данных', f'Файл успешно сохранен! Добавлено строк: {len(data)}')
                clear_data()
            else:
                error("Ошибка входных данных", "Введите требуемые значения!")

    def del_city():
        que = messagebox.askokcancel(title="Удалить город из списка", message=f"Вы уверены, что хотите удалить {combo_city.get()} из списка?")
        if que:
            path_dir_txt = Path(user_path + '/Города')
            file_path = Path(rf"{path_dir_txt}\cities.txt")
            with open(file_path) as f:
                lines = f.readlines()

            str = combo_city.get()
            pattern = re.compile(re.escape(str))
            with open(file_path, 'w') as f:
                for line in lines:
                    result = pattern.search(line)
                    if result is None:
                        f.write(line)

            upd_cities = r_cities_txt()
            combo_city.config(values=upd_cities)
            combo_city.current(2)


    def clear_data():
        # e_city.delete(0, END)
        e_fio.delete(0, END)
        e_director.delete(0, END)
        e_doc.delete(0, END)
        e_job.delete(0, END)
        e_name.delete(0, END)
        e_seria.delete(0, END)
        e_number.delete(0, END)
        e_point_of_get.delete(0, END)
        e_directors_job.delete(0, END)
        e_directors_name.delete(0, END)
        e_date.delete(0, END)

        e_code.delete(0, END)
        e_birth.delete(0, END)
        e_place_of_birth.delete(0, END)
        # e_gender.delete(0, END)

        e_full_fio.delete(0, END)
        e_inn.delete(0, END)
        e_snils.delete(0, END)
        e_mail.delete(0, END)

    def select():
        nonlocal gender_data
        l = gender.get()
        if l == 1:
            gender_data = 'М'
        elif l == 2:
            gender_data = 'Ж'

    def on_closing():
        win.destroy()
        root.deiconify()

    def close_ad(city):
        w_cities_txt(city)
        upd_cities = r_cities_txt()
        combo_city.config(values=upd_cities)
        combo_city.current(len(upd_cities)-1)

    def r_cities_txt():
        path_dir_txt = Path(user_path + '/Города')
        if not path_dir_txt.exists():
            os.mkdir(path_dir_txt)

        file_path = Path(rf"{path_dir_txt}\cities.txt")
        if not file_path.exists():
            my_file = open(rf"{path_dir_txt}\cities.txt", "w+")
            for city in cities:
                my_file.write(city + f'\n')
            my_file.close()

        data = []
        r_file = open(rf"{path_dir_txt}\cities.txt", "r").readlines()
        for line in r_file:
            data.append(line)
        return data

    def w_cities_txt(temp):
        path_dir_txt = Path(user_path + '/Города')
        if not path_dir_txt.exists():
            os.mkdir(path_dir_txt)

        file_path = Path(rf"{path_dir_txt}\cities.txt")
        if not file_path.exists():
            my_file = open(rf"{path_dir_txt}\cities.txt", "w+")
            for city in cities:
                my_file.write(city + f'\n')
            my_file.write(temp + f'\n')
            my_file.close()
        else:
            my_file = open(rf"{path_dir_txt}\cities.txt", "a+")
            my_file.write(temp + f'\n')
            my_file.close()


    def update_cities():
        def close_win_ad():
            ad.destroy()
            win.deiconify()

        def add():
            city = e_new_city.get()
            if city != '':
                close_ad(city)
                ad.destroy()
                win.deiconify()
            else:
                error("Ошибка ввода", "Вы ввели некорректное значение!")

        ad = Toplevel()
        ad.geometry('400x100+100+50')
        ad.title('Введите новый город')
        ad.grab_set()
        win.withdraw()
        ad.protocol("WM_DELETE_WINDOW", close_win_ad)
        ad.resizable(False, False)
        ad.config(bg="#F1EEE9")

        f0 = Frame(ad, bg="#F1EEE9")
        f0.pack(fill=BOTH, padx=10, pady=10)

        e_new_city = Entry(f0)
        e_new_city.pack(fill=X, padx=10, ipady=2, expand=True)

        btn_add_city = ttk.Button(f0, text="Принять", command=add)
        btn_add_city.pack(side=BOTTOM, padx=5)

    def on_mousewheel1(event):
        my_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def set_binds_canvas1(event):
        win.bind_all("<MouseWheel>", on_mousewheel1)

    win = Toplevel()
    win.geometry('1500x590+100+50')
    win.title('Внесение данных в таблицу')
    win.grab_set()
    root.withdraw()
    win.protocol("WM_DELETE_WINDOW", on_closing)
    # win.resizable(True, False)
    # win.minsize(1000, 550)
    win.config()


    # canvas, который будет прокручиваться sb
    my_canvas = Canvas(win, bg='green')
    my_canvas.pack(side=LEFT, fill=BOTH, expand=True)

    # sec_f, который создаётся методом от canvas и в нём будут лежать фреймы
    sec_f = Frame(my_canvas, bg="blue")
    sec_f.pack(side=LEFT, fill=BOTH, expand=True)

    # scrollbar
    my_scrollbar = ttk.Scrollbar(win, orient=VERTICAL, command=my_canvas.yview)
    my_scrollbar.pack(side=RIGHT, fill=Y)

    my_canvas.configure(yscrollcommand=my_scrollbar.set)
    my_canvas.bind(
        '<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all"))
    )

    my_canvas.create_window((0, 0), window=sec_f)
    win.bind("<Enter>", set_binds_canvas1)

    # f0 = Frame(my_canvas, bg="#F1EEE9")
    # f0.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW, expand=True)
    #
    # f1 = Frame(my_canvas, bg="#F1EEE9")
    # f1.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW, expand=True)
    #
    # f2 = Frame(my_canvas, bg="#F1EEE9")
    # f2.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f3 = Frame(my_canvas, bg="#F1EEE9")
    # f3.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f3_1 = Frame(my_canvas, bg="#F1EEE9")
    # f3_1.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f4 = Frame(my_canvas, bg="#F1EEE9")
    # f4.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f5 = Frame(my_canvas, bg="#F1EEE9")
    # f5.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f6 = Frame(my_canvas, bg="#F1EEE9")
    # f6.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f7 = Frame(my_canvas, bg="#F1EEE9")
    # f7.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f8 = Frame(my_canvas, bg="#F1EEE9")
    # f8.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f9 = Frame(my_canvas, bg="#F1EEE9")
    # f9.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f10 = Frame(my_canvas, bg="#F1EEE9")
    # f10.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f11 = Frame(my_canvas, bg="#F1EEE9")
    # f11.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f12 = Frame(my_canvas, bg="#F1EEE9")
    # f12.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NW)
    #
    # f13 = Frame(my_canvas, bg="#F1EEE9")
    # f13.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f14 = Frame(my_canvas, bg="#F1EEE9")
    # f14.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f15 = Frame(sec_f, bg="#F1EEE9")
    # f15.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f16 = Frame(sec_f, bg="#F1EEE9")
    # f16.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f17 = Frame(sec_f, bg="#F1EEE9")
    # f17.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f18 = Frame(sec_f, bg="#F1EEE9")
    # f18.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f19 = Frame(sec_f, bg="#F1EEE9")
    # f19.pack(side=TOP, fill=X, padx=10, pady=10, anchor=NE)
    #
    # f20 = Frame(sec_f, bg="#F1EEE9")
    # f20.pack(side=TOP, fill=X, padx=10, pady=10, anchor=S)
    #
    # f21 = Frame(sec_f, bg="#F1EEE9")
    # f21.pack(side=TOP, fill=X, padx=10, pady=10, anchor=S)

    label_0 = Label(sec_f, width=25, text='Город:')
    label_0.pack(side=LEFT)

    btn_add_city = ttk.Button(sec_f, text="Другой", command=update_cities)
    btn_add_city.pack(side=LEFT, padx=5)

    cities = ['Ачинск', 'Канск', 'Красноярск', 'Лесосибирск', 'Минусинск', 'Норильск']
    actual_cities = r_cities_txt()
    combo_city = ttk.Combobox(sec_f, values=actual_cities)
    combo_city.current(2)
    combo_city.pack(side=LEFT, fill=X, padx=10)

    btn_del_city = ttk.Button(sec_f, text="<- Удалить", command=del_city)
    btn_del_city.pack(side=LEFT, padx=2)

    label_1 = Label(sec_f, width=25, text='Фамилия И.О.:')
    label_1.pack(side=TOP, anchor=NW, expand=True)

    e_fio = Entry(sec_f)
    e_fio.pack(side=TOP, anchor=NW, fill=X, padx=10, ipady=2, expand=True)

    label_2 = Label(sec_f, width=25, text='В лице руководителя:')
    label_2.pack(side=TOP, anchor=NW )

    e_director = Entry(sec_f)
    e_director.pack(side=TOP, anchor=NW, fill=X, padx=10, ipady=2, expand=True)

    label_3 = Label(sec_f, width=25, text='На основании:')
    label_3.pack(side=TOP, anchor=NW)

    e_doc = Entry(sec_f)
    e_doc.pack(side=TOP, anchor=NW, fill=X, padx=10, ipady=2, expand=True)

    label_3_1 = Label(sec_f, width=25, text='Уполномочивает', background='grey', foreground='yellow')
    label_3_1.pack(side=LEFT)

    label_4 = Label(sec_f, width=25, text='Должность:', background='grey', foreground='yellow')
    label_4.pack(side=LEFT)

    e_job = Entry(sec_f)
    e_job.pack(side=TOP, anchor=NW, fill=X, padx=10, ipady=2, expand=True)

    label_5 = Label(sec_f, width=25, text='Сотрудника (полн.ФИО):', background='grey', foreground='yellow')
    label_5.pack(side=LEFT)

    e_name = Entry(sec_f)
    e_name.pack(fill=X, padx=10, ipady=2, expand=True)

    label_17 = Label(sec_f, width=25, text='ИНН:')
    label_17.pack(side=LEFT)

    e_inn = Entry(sec_f)
    e_inn.pack(fill=X, padx=10, ipady=2, expand=True)

    label_18 = Label(sec_f, width=25, text='СНИЛС:')
    label_18.pack(side=LEFT)

    e_snils = Entry(sec_f)
    e_snils.pack(fill=X, padx=10, ipady=2, expand=True)

    label_6 = Label(sec_f, width=25, text='Серия паспорта:')
    label_6.pack(side=LEFT)

    e_seria = Entry(sec_f)
    e_seria.pack(fill=X, padx=10, ipady=2, expand=True)

    label_7 = Label(sec_f, width=25, text='Номер паспорта:')
    label_7.pack(side=LEFT)

    e_number = Entry(sec_f)
    e_number.pack(fill=X, padx=10, ipady=2, expand=True)

    label_9 = Label(sec_f, width=25, text='Кем выдан:')
    label_9.pack(side=LEFT)

    e_point_of_get = Entry(sec_f)
    e_point_of_get.pack(fill=X, padx=10, ipady=2, expand=True)

    label_8 = Label(sec_f, width=25, text='Дата выдачи паспорта:')
    label_8.pack(side=LEFT)

    e_date = Entry(sec_f)
    e_date.pack(fill=X, padx=10, ipady=2, expand=True)

    label_12 = Label(sec_f, width=25, text='Код места выдачи:')
    label_12.pack(side=LEFT)

    e_code = Entry(sec_f)
    e_code.pack(fill=X, padx=10, ipady=2, expand=True)

    label_16 = Label(sec_f, width=25, text='ФИО (полное):')
    label_16.pack(side=LEFT)

    e_full_fio = Entry(sec_f)
    e_full_fio.pack(fill=X, padx=10, ipady=2, expand=True)

    label_15 = Label(sec_f, width=25, text='Пол:')
    label_15.pack(side=LEFT)

    gender = IntVar()

    male_checkbutton = Radiobutton(sec_f, text="М", value=1, variable=gender, padx=10, command=select)
    male_checkbutton.pack(side=LEFT)

    female_checkbutton = Radiobutton(sec_f, text="Ж", value=2, variable=gender, padx=10, command=select)
    female_checkbutton.pack(side=LEFT)

    label_13 = Label(sec_f, width=25, text='Дата рождения:')
    label_13.pack(side=LEFT)

    e_birth = Entry(sec_f)
    e_birth.pack(fill=X, padx=10, ipady=2, expand=True)

    label_14 = Label(sec_f, width=25, text='Место рождения:')
    label_14.pack(side=LEFT)

    e_place_of_birth = Entry(sec_f)
    e_place_of_birth.pack(fill=X, padx=10, ipady=2, expand=True)

    label_10 = Label(sec_f, width=25, text='Должность руководителя:')
    label_10.pack(side=LEFT)

    e_directors_job = Entry(sec_f)
    e_directors_job.pack(fill=X, padx=10, ipady=2, expand=True)

    label_11 = Label(sec_f, width=25, text='И.О.Фамилия руководителя:')
    label_11.pack(side=LEFT)

    e_directors_name = Entry(sec_f)
    e_directors_name.pack(fill=X, padx=10, ipady=2, expand=True)

    label_19 = Label(sec_f, width=25, text='Почта:')
    label_19.pack(side=LEFT)

    e_mail = Entry(sec_f)
    e_mail.pack(fill=X, padx=10, ipady=2, expand=True)

    btn_confirm = ttk.Button(sec_f, text="Принять", command=confirm_data)
    btn_confirm.pack(side=RIGHT, padx=10)

    btn_clear_e = ttk.Button(sec_f, text="Очистить", command=clear_data)
    btn_clear_e.pack(side=LEFT, padx=5)

    btn_fill_e = ttk.Button(sec_f, text="Заполнить", command=insert_data)
    btn_fill_e.pack(side=LEFT, padx=5)

    btn_fill_e = ttk.Button(sec_f, text="Шаблон", command=open_xlsx_sample)
    btn_fill_e.pack(side=LEFT, padx=17)

    label12 = ttk.Label(sec_f, text="Статус загрузки файлов:")
    label12.pack(side=LEFT, padx=10)

    label13 = ttk.Label(sec_f, text="Таблица с данными", foreground='red', justify=LEFT)
    label13.pack(side=RIGHT, padx=10)

    main_menu = Menu(win)
    win.config(menu=main_menu)



    # Выбрать таблицу

    file_menu = Menu(main_menu, tearoff=0)
    file_menu.add_command(label="Выбрать таблицу c данными", command=open_xlsx)
    file_menu.add_command(label="Очистить таблицу", command=clear_xlsx)
    file_menu.add_command(label="Создать таблицу", command=create_xlsx)
    main_menu.add_cascade(label="Таблица", menu=file_menu, )

    insert_data()


root = Tk()
root.title('Доверенность')
root.geometry('325x150+100+50')
root.resizable(False, False)

main_menu = Menu(root)
root.config(menu=main_menu)

# Автоматический ввод данных
file_menu = Menu(main_menu, tearoff=0)
file_menu.add_command(label="Выбрать таблицу c данными", command=open_table)
file_menu.add_command(label="Выбрать шаблон", command=open_sample_doc)
main_menu.add_cascade(label="Шаблон", menu=file_menu, )

# Рцчной ввод данных
fill_menu = Menu(main_menu, tearoff=0)
# fill_menu.add_command(label="Открыть таблицу", command=open_table)
fill_menu.add_command(label="Заполнить вручную", command=open_custom_fill)
main_menu.add_cascade(label="Данные", menu=fill_menu, )

f1 = Frame(root)
f1.pack(fill=X, padx=10, pady=5)

f2 = Frame(root)
f2.pack(fill=X, padx=10, pady=5)

f3 = Frame(root)
f3.pack(fill=X, padx=10, pady=5)

f4 = Frame(root)
f4.pack(fill=X, padx=10, pady=5)

label1 = ttk.Label(f1, text="Начальная точка:", width=17)
label1.pack(side=LEFT, padx=10)

e_start = ttk.Entry(f1)
e_start.pack(side=RIGHT, fill=X, expand=True)

label2 = ttk.Label(f2, text="Конечная точка:", width=17)
label2.pack(side=LEFT, padx=10)

e_end = ttk.Entry(f2)
e_end.pack(side=RIGHT, fill=X, expand=True)

btn_show_table = ttk.Button(f3, text="Заполнить", command=fill_files)
btn_show_table.pack(side=RIGHT)

btn_show_dir = ttk.Button(f3, text="Результат", command=open_dir)
btn_show_dir.pack(side=LEFT, padx=10)

label3_1 = ttk.Label(f4, text="Статус загрузки файлов: ")
label3_1.pack(side=LEFT, padx=10)

label3 = ttk.Label(f4, text="Шаблоны", foreground='red', justify=LEFT)
label3.pack(side=BOTTOM, padx=10)

label4 = ttk.Label(f4, text="Таблица с данными", foreground='red', justify=LEFT)
label4.pack(side=BOTTOM, padx=10)

e_start.insert(0, 'B2')

root.mainloop()
