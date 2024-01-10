import openpyxl
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from validate_email import validate_email
from imap_tools import MailBox
import html2text
import messages as mes
import variables as var
import check_funcs as ch_foo
import service


def requests_win(sh):
    def show_templs():
        # path_table_temples = ''

        def templs_closing():
            templs.destroy()
            mail.deiconify()

        def show_table():
            def clear_e_templs():
                e_out_fio.delete(0, END)
                e_out_code.delete(0, END)
                e_out_url.delete(0, END)

            def ins_e_templs(record):
                t_fio = record[1]
                t_code = record[3]
                t_url = record[4]
                e_out_fio.insert(0, t_fio)
                e_out_code.insert(0, t_code)
                e_out_url.insert(0, t_url)

            def temp_selected(event):
                for selected_item in table.selection():
                    item = table.item(selected_item)
                    record = item['values']
                    clear_e_templs()
                    ins_e_templs(record)

            if var.path_table != '':
                for widget in frame_4.winfo_children():
                    widget.destroy()

                heads = ['№', 'ФИО', 'Дата', 'Код', 'Ссылка']
                lst = get_data_temples()

                table = ttk.Treeview(frame_4, show='headings')
                table['columns'] = heads

                for header in heads:
                    table.heading(header, text=header, anchor=CENTER)
                    table.column(header, anchor=CENTER, width=1)
                for row in lst:
                    table.insert('', END, values=row)

                scroll = ttk.Scrollbar(frame_4, command=table.yview)
                table.configure(yscrollcommand=scroll.set)
                scroll.pack(side=RIGHT, fill=Y)

                table.bind('<<TreeviewSelect>>', temp_selected)

                table.pack(expand=True, side=TOP, padx=5, fill=BOTH)
            else:
                mes.error("Таблица с данными", "Выберите таблицу для обработки!")

        def get_data_temples():
            if var.path_table == '':
                mes.error("Таблица с данными", "Выберите таблицу для обработки!")
            else:
                wb = openpyxl.load_workbook(var.path_table)
                sheet = wb["Черновики"]

                start = 'A2'
                end = sheet['D'][-1].coordinate
                lst = []
                test = []
                rows = 0

                try:
                    for row in sheet[f'{start}':f'{end}']:
                        rows += 1
                        for cellObj in row:
                            if cellObj.value == None or cellObj.value == '':
                                continue

                            test.append(str(cellObj.value))
                    x = 0
                    n = 1
                    while x < len(test):
                        lst.append((str(n), test[x], test[x + 1], test[x + 2], test[x + 3]))
                        n += 1
                        x += 4  # переход на новую строку
                    wb.close()
                    return lst
                except:
                    mes.error("Обработка входных данных", "Ошибка обработки данных!")
                    wb.close()

        # def open_tab():
        #     # nonlocal path_table_temples
        #     path_table_temples = filedialog.askopenfilename(title="Выбор таблицы для заполнения",
        #                                                     initialdir=var.user_docs_path,
        #                                                     filetypes=(("Таблицы (*.xlsx)",
        #                                                                 "*.xlsx"),
        #                                                                ("Все файлы", "*.*")))

        templs = Toplevel()
        templs.geometry('550x400+100+50')
        templs.title('Вывод значений')
        templs.minsize(550, 400)
        templs.grab_set()
        mail.withdraw()
        templs.protocol("WM_DELETE_WINDOW", templs_closing)
        templs.resizable(True, False)
        templs.config(bg="#F1EEE9")

        frame_1 = Frame(templs, bg="#F1EEE9")
        frame_1.pack(fill=X, padx=5)

        frame_2 = Frame(templs, bg="#F1EEE9")
        frame_2.pack(fill=X, padx=5)

        frame_3 = Frame(templs, bg="#F1EEE9")
        frame_3.pack(fill=X, padx=5)

        frame_4 = Frame(templs, bg="#F1EEE9")
        frame_4.pack(fill=BOTH, padx=5, pady=5, ipady=5, ipadx=5)

        labe_1 = ttk.Label(frame_1, text="Для вывода данных выберите субъекта в таблице!")
        labe_1.pack(side=LEFT, padx=5)

        e_out_fio = ttk.Entry(frame_2)
        e_out_fio.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

        e_out_code = ttk.Entry(frame_2)
        e_out_code.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

        e_out_url = ttk.Entry(frame_2)
        e_out_url.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

        btn_show_data = ttk.Button(frame_3, text="Показать таблицу", command=show_table)
        btn_show_data.pack(side=RIGHT, padx=5, pady=5, ipadx=2)

        m_menu = Menu(templs)
        templs.config(menu=m_menu)

        # Выбрать таблицу
        in_menu = Menu(m_menu, tearoff=0)
        in_menu.add_command(label="Указать таблицу c данными", command=service.open_table)
        m_menu.add_cascade(label="Таблица", menu=in_menu, )

    def mail_closing():
        mail.destroy()
        sh.deiconify()

    def check_email(mail):
        check = mail
        is_valid = validate_email(check, verify=True)
        if is_valid:
            return True
        else:
            return False

    def get_pass_code():
        if e_login.get() != '':
            if e_imap.get() != '':
                var.mail_imap = e_imap.get()
                if e_login.get() == 'ikdomashenko@kkck.ru':
                    var.login = e_login.get()
                    return True
                else:
                    if check_email(e_login.get()):
                        print(e_login.get(), e_pass.get(), 'Укажите код' in e_pass.get())
                        if not 'Укажите код' in e_pass.get() and e_imap.get() != '':
                            if e_pass.get() != '':
                                var.login = e_login.get()
                                return True
                            else:
                                mes.error('Ошибка авторизации', 'Укажите код доступа к вашей почте для сторонних приложений imap!')
                                return False
                        else:
                            mes.error('Ошибка авторизации',
                                      'Укажите код доступа к вашей почте для сторонних приложений imap и порт imap вашей почты!')
                            return False
                    else:
                        mes.error('Ошибка', 'Введите корректный адрес электронной почты для обработки!')
                        return False
            else:
                mes.error('Ошибка', 'Введите корректное значение imap сервера!')
                return False
        else:
            mes.error('Ошибка', 'Введите адрес электронной почты для обработки!')
            return False

    def update_table(send):
        def insert_templates(send):
            if var.path_table == '':
                mes.error("Ошибка входных данных", "Укажите таблицу данных!")
            else:
                wb = openpyxl.load_workbook(var.path_table)
                sheet = wb["Черновики"]
                data = send
                nb_row = sheet.max_row
                if nb_row == 1:
                    for row in data:
                        sheet.append(row)
                else:
                    for row in data:
                        start_name = 'B2'
                        end_name = sheet['B'][-1].coordinate
                        for item in sheet[f'{start_name}':f'{end_name}']:
                            sheet.append(row)
                wb.save(var.path_table)
                mes.info('Обработка данных', f'Файл успешно сохранен! Добавлено строк: {len(data)}')

        insert_templates(send)

    def get_login_data():
        if var.path_table == '':
            mes.error("Ошибка входных данных", "Укажите таблицу данных!")
        else:
            curent = 0
            labe2['text'] = f"Проверяем учетные данные из формы..."
            progress_bar['value'] = 0
            frame4.update()

            send = list()
            new_mails = list()

            if check_email(e_login.get()):
                if get_pass_code():
                    labe2['text'] = f"Соединяемся с сервером..."
                    frame4.update()
                    try:
                        login = var.login
                        password = e_pass.get()
                        if password.find('Укажите код') == -1:
                            if password != '':
                                var.mail_code = password
                                print(var.mail_code)

                        print(var.mail_code)
                        mes.warning(title='Предупреждение', text='Внимание!\nПрограмма может не '
                                                                               'отвечать после начала операции '
                                                                               'из-за подключения к почтовому '
                                                                               'серверу.\n\nПожалуйста, '
                                                                               'дождитесь уведомления о '
                                                                               'завершении процесса!\nВремя '
                                                                               'выполнения зависит от ресурсов '
                                                                               'вашего компьютера (~1 min.)')

                        from datetime import datetime, timedelta
                        check_date = datetime.now().date() - timedelta(days=35)
                        print(datetime.now().date(), check_date)

                        labe2['text'] = f"Просматриваем письма..."
                        frame4.update()

                        with MailBox(var.mail_imap).login(login, var.mail_code) as mailbox:
                            mailbox.folder.set('INBOX/Казначейство/Запросы/Черновик')
                            mes_limit_str = ''
                            for char in str(mailbox.folder.set('INBOX/Казначейство/Запросы/Черновик')[1][0]):
                                if char.isdigit():
                                    mes_limit_str += char
                            for msg in mailbox.fetch(limit=int(mes_limit_str), mark_seen=False, miss_no_uid=False, bulk=True, reverse=False):
                                if msg.from_ == 'uc_fk@roskazna.ru' and msg.subject == 'Создан черновик запроса на сертификат':
                                    print(f'Смотрю {msg.date} - {msg.date.date()} - {check_date} - {msg.date.date() >= check_date}, {msg.subject}')
                                    if msg.date.date() >= check_date:
                                        print(f'Нашел {msg.date.date()}, {msg.subject}')
                                        from_ = msg.from_
                                        subject = msg.subject
                                        date = msg.date
                                        body = msg.text or msg.html
                                        text = html2text.html2text(body)
                                        res = text.replace('\n', '')
                                        index_code = res.find('Номер запроса: ') + 15
                                        index_fio = res.find('ФИО: ') + 5
                                        index_fio_end = res.find('  Черновик запроса будет доступен по ссылке до')
                                        index_url = res.find('[ссылка]') + 9
                                        index_url_end = res.find(').  **Сведения запроса')

                                        new_mails.append([res[index_fio:index_fio_end],
                                                        date.strftime('%d.%m.%Y'),
                                                        res[index_code:index_code + 6],
                                                        res[index_url:index_url_end]])

                                        # Проверяем наличие запроса
                                        wb = openpyxl.load_workbook(var.path_table)
                                        sheet = wb["Черновики"]
                                        exists_flag = False
                                        for row in sheet.values:
                                            print(res[index_fio:index_fio_end], row[0])
                                            print(res[index_code:index_code + 6], row[2])
                                            if res[index_fio:index_fio_end] == row[0] and res[index_code:index_code + 6] == row[2]:
                                                exists_flag = True
                                                break
                                            else:
                                                continue
                                        wb.close()

                                        for row in send:
                                            if res[index_fio:index_fio_end] == row[0] and res[index_code:index_code + 6] == row[2]:
                                                exists_flag = True
                                                break
                                            else:
                                                continue

                                        # Если не существует, то записываем
                                        if not exists_flag:
                                            temp = [res[index_fio:index_fio_end],
                                                    date.strftime('%d.%m.%Y'),
                                                    res[index_code:index_code + 6],
                                                    res[index_url:index_url_end]]
                                            send.append(temp)
                                            progress_bar['value'] += 1
                                            curent += 1
                                            labe2['text'] = f"Обработано строк: {curent}"
                                            frame4.update()
                                    else:
                                        continue
                                else:
                                    continue
                            progress_bar['value'] = 100
                            frame4.update()
                            mes.info('Обработка почты', 'Обработка завершена, переходим к записи данных в таблицу!')
                            if len(send) > 1:
                                labe2['text'] = f"Записываем в таблицу..."
                                frame4.update()
                                print(len(send), send)
                                update_table(send)

                                labe2['text'] = f"Записано в файл строк: {len(send)}"
                                frame4.update()

                            else:
                                labe2['text'] = f"Найдено актуальных запросов: 0"
                                frame4.update()
                                mes.info('Обработка запросов с почты', 'Не найдены актуальные запросы!')
                    except:
                        labe2['text'] = f"Ошибка обработки"
                        progress_bar['value'] = 0
                        frame4.update()
                        mes.error('Обработка почтового сервера', 'При обработке возникла непредвиденная ошибка!')
                else:
                    mes.error('Ошибка', 'Некорректные данные для авторизации!')
                    labe2['text'] = f"Ошибка авторизации"
                    progress_bar['value'] = 0
                    frame4.update()
            else:
                messagebox.showerror(title="Проверка почты", message="Почта не существует!")
                labe2['text'] = f"Ошибка авторизации"
                progress_bar['value'] = 0
                frame4.update()

    def clear_requests():
        if var.path_table != '':
            question = messagebox.askokcancel(title="Очистка черновиков", message="Вы уверены, что хотите очистить "
                                                                                  "записанные черновики?\nВсе "
                                                                                  "данные в листе будут безвозвратно "
                                                                                  "удалены!")
            if question:
                try:
                    book = openpyxl.load_workbook(var.path_table)
                    sheet3 = book["Черновики"]

                    last_row3 = len(list(sheet3.rows))
                    first_row3 = 2
                    sheet3.delete_rows(first_row3, last_row3)

                    del3 = last_row3 - first_row3

                    if del3 < 0:
                        del3 = 0
                    book.save(var.path_table)
                    mes.info("Очистка таблицы", f'Успешно удалено строк:\nВ листе "Черновики: "{del3}')
                except:
                    mes.error('Ошибка', 'Лист не найден или таблица уже открыта!')
        else:
            mes.error("Ошибка входных данных", "Выберите таблицу с данными!")

    def get_table():
        if service.open_table():
            label2['foreground'] = 'green'
        else:
            label2['foreground'] = 'red'

    def full_exit():
        service.full_exit(mail)

    mail = Toplevel()
    mail.geometry('500x250+100+50')
    mail.title('Получение черновиков')
    mail.grab_set()
    sh.withdraw()
    mail.protocol("WM_DELETE_WINDOW", mail_closing)
    mail.resizable(False, False)
    mail.config(bg="#F1EEE9")

    frame7 = Frame(mail, bg="#F1EEE9")
    frame7.pack(fill=X, padx=5, pady=5)

    frame1 = Frame(mail, bg="#F1EEE9")
    frame1.pack(fill=X, padx=5)

    frame2 = Frame(mail, bg="#F1EEE9")
    frame2.pack(fill=X, padx=5)

    frame3 = Frame(mail, bg="#F1EEE9")
    frame3.pack(fill=X, padx=5)

    frame3_1 = Frame(mail, bg="#F1EEE9")
    frame3_1.pack(fill=X, padx=5)

    frame4 = Frame(mail, bg="#F1EEE9")
    frame4.pack(fill=X, padx=5, pady=10)

    frame5 = Frame(mail, bg="#F1EEE9")
    frame5.pack(fill=X, padx=5)

    frame6 = Frame(mail, bg="#F1EEE9")
    frame6.pack(fill=X, padx=5)

    labe1 = ttk.Label(frame1, text="Укажите почту: ")
    labe1.pack(side=LEFT, padx=5)

    e_login = ttk.Entry(frame2)
    e_login.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)
    e_login.delete(0, END)
    e_login.insert(0, 'ikdomashenko@kkck.ru')

    e_pass = ttk.Entry(frame3)
    e_pass.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)
    e_pass.insert(0, 'Укажите код для доступа к приложению, если нужна другая почта')

    e_imap = ttk.Entry(frame3_1)
    e_imap.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)
    e_imap.insert(0, var.mail_imap)

    labe2 = ttk.Label(frame4, text=f"Запустите обработку запросов")
    labe2.pack(side=LEFT, padx=5)

    progress_bar = ttk.Progressbar(frame4, orient="horizontal", mode="determinate",
                                   maximum=100, value=0)
    progress_bar.pack(side=LEFT, fill=X, padx=10, ipady=2, expand=True, pady=5)

    btn_send = ttk.Button(frame6, text="Получить запросы", command=get_login_data)
    btn_send.pack(side=LEFT, padx=5)

    btn_show_templs = ttk.Button(frame6, text="Показать черновики", command=show_templs)
    btn_show_templs.pack(side=LEFT, padx=5)

    btn_temp = ttk.Button(frame6, text="Шаблоны", command=ch_foo.open_xlsx_sample)
    btn_temp.pack(side=RIGHT, padx=5)

    btn_full_exit = ttk.Button(frame7, text="Выйти из программы", command=full_exit)
    btn_full_exit.pack(side=RIGHT, padx=5)

    label2 = ttk.Label(frame7, text="Таблица с данными", foreground='red')
    label2.pack(side=LEFT, padx=5)

    if var.path_table != '':
        label2['foreground'] = 'green'
    else:
        label2['foreground'] = 'red'

    req_menu = Menu(sh)
    mail.config(menu=req_menu)

    # Выбрать таблицу
    tab_menu = Menu(req_menu, tearoff=0)
    tab_menu.add_command(label="Указать таблицу c данными", command=get_table)
    tab_menu.add_command(label="Очистить черновики", command=clear_requests)
    req_menu.add_cascade(label="Таблица", menu=tab_menu, )