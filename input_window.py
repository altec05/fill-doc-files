import openpyxl
from tkinter import *
from tkinter import messagebox
import os
from pathlib import Path
from tkinter import ttk
import re
import messages as mes
import output_window
import variables as var
import check_funcs as ch_foo
import service
import get_requests_window


def open_custom_fill(root):
    def on_closing():
        win.destroy()
        root.deiconify()

    def check_entry():
        if gender_data != '' and e_fio.get() and e_director.get() and e_doc.get() and \
                e_job.get() and e_name.get() and e_seria.get() and e_number.get() and e_date.get() and \
                e_point_of_get.get() and e_directors_job.get() and e_directors_name.get() and \
                e_code.get() and e_birth.get() and e_place_of_birth.get() and gender.get() and e_full_fio.get() and \
                e_inn.get() and e_snils.get() and e_mail.get():
            return True
        else:
            return False

    def create_xlsx():
        result = True
        if ch_foo.check_table():
            result = mes.ask('Создание таблицы', 'Внимание!\n\nТаблица уже существует!\n\nПерезаписать таблицу?')

        if result:
            book = openpyxl.Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("Доверенность")
            sheet_2 = book.create_sheet("Заявление")
            sheet_3 = book.create_sheet("Черновики")

            sheet_1.insert_rows(0)
            col_names = ['№', 'Город', 'На кого', 'Руководитель', 'Основание', 'Должность', 'ФИО', 'Серия паспорта',
                         'Номер паспорта', 'Дата выдачи', 'Кем выдан', 'Должность руководителя', 'ФИО руководителя']
            l = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                 'V', 'W', 'X', 'Y', 'Z']
            n = 0
            for i in col_names:
                sheet_1[f'{l[n]}1'].value = f'{col_names[n]}'
                sheet_1.column_dimensions[l[n]].width = 20
                n += 1

            sheet_2.insert_rows(0)
            col2_names = ['№', 'Серия', 'Номер', 'Дата выдачи', 'Код', 'Дата рождения', 'Место рождения', 'Пол',
                          'Фамилия', 'Имя', 'Отчество', 'ИНН', 'СНИЛС', 'Должность', 'Почта', 'Населенный пункт']
            k = 0
            for i in col2_names:
                sheet_2[f'{l[k]}1'].value = f'{col2_names[k]}'
                sheet_2.column_dimensions[l[k]].width = 20
                k += 1

            sheet_3.insert_rows(0)
            col3_names = ['ФИО', 'Дата', 'Код', 'Ссылка']
            m = 0
            for i in col3_names:
                sheet_3[f'{l[m]}1'].value = f'{col3_names[m]}'
                sheet_3.column_dimensions[l[m]].width = 20
                m += 1

            path = Path(var.sample_folder_path)
            if not path.exists():
                os.makedirs(path)

            book.save(var.table_xlsx_file_path)
            mes.info("Создание таблицы", 'Таблица успешно создана!')
        else:
            mes.warning('Создание таблицы', 'Операция отменена пользователем!')

    def clear_xlsx():
        if var.path_table != '':
            question = messagebox.askokcancel(title="Очистка таблицы", message="Вы уверены, что хотите очистить "
                                                                               "таблицу?\nВсе "
                                                                               "данные в таблице будут безвозвратно удалены!")
            if question:
                book = openpyxl.load_workbook(var.path_table)
                sheet = book["Доверенность"]
                sheet2 = book["Заявление"]
                last_row = len(list(sheet.rows))
                first_row = 1
                sheet.delete_rows(first_row, last_row)

                last_row2 = len(list(sheet2.rows))
                first_row2 = 1
                sheet2.delete_rows(first_row2, last_row2)

                del1 = last_row - first_row
                del2 = last_row2 - first_row2
                if del1 < 0:
                    del1 = 0
                if del2 < 0:
                    del2 = 0
                book.save(var.path_table)
                mes.info("Очистка таблицы", f'Успешно удалено строк: \nВ листе "Доверенность" - {del1}\n'
                                        f'В листе "Заявление" - {del2}')
        else:
            mes.error("Ошибка входных данных", "Выберите таблицу с данными!")

    gender_data = ''

    def insert_data():
        clear_for_fill()
        e_director.insert(0, 'Главного врача Филиной Натальи Григорьевны')
        e_doc.insert(0, 'Устава, утвержденного Приказом от 28.12.2021 №2586-орг')
        e_directors_job.insert(0, 'Главный врач')
        e_directors_name.insert(0, 'Н.Г. Филина')
        e_mail.insert(0, 'ikdomashenko@kkck.ru')

    def get_data():
        city = combo_city.get().rstrip()
        data = [[city, e_fio.get(), e_director.get(), e_doc.get(), e_job.get(), e_name.get(), e_seria.get(),
                 e_number.get(), e_date.get(), e_point_of_get.get(), e_directors_job.get(), e_directors_name.get()]]
        return data

    def get_data2():
        temp = e_full_fio.get()
        temp2 = temp.split()
        if len(temp2) < 3:
            mes.error("Ошибка входных данных", "Укажите верное ФИО (полное)!")
        else:
            fam = temp2[0]
            name = temp2[1]
            otch = temp2[2]
            temp = e_snils.get()
            snils = "".join(c for c in temp if c.isalnum())
            city = combo_city.get().rstrip()
            data2 = [
                [
                    e_seria.get(), e_number.get(), e_date.get(), e_code.get(), e_birth.get(), e_place_of_birth.get(),
                    gender_data, fam, name, otch, e_inn.get(), snils, e_job.get(), e_mail.get(), city
                ]
            ]
            return data2

    def confirm_data():
        if var.path_table == '':
            mes.error("Ошибка входных данных", "Укажите таблицу данных!")
        else:
            if check_entry():
                wb = openpyxl.load_workbook(var.path_table)
                sheet = wb["Доверенность"]
                temp = get_data()
                try:
                    number = int(sheet["A"][-1].value) + 1
                except:
                    number = 1
                data = temp
                data[0].insert(0, number)
                for row in data:
                    sheet.append(row)

                sheet2 = wb["Заявление"]
                temp2 = get_data2()
                try:
                    number = int(sheet2["A"][-1].value) + 1
                except:
                    number = 1
                data2 = temp2
                data2[0].insert(0, number)
                for row in data2:
                    sheet2.append(row)
                wb.save(var.path_table)
                mes.info('Обработка данных', f'Файл успешно сохранен! Добавлено строк: {len(data)}')
                clear_data()
            else:
                mes.error("Ошибка входных данных", "Введите требуемые значения!")


    def del_city():
        que = messagebox.askokcancel(title="Удалить город из списка",
                                     message=f"Вы уверены, что хотите удалить {combo_city.get()} из списка?")
        if que:
            path_dir_txt = Path(var.cities_path)
            file_path = Path(rf"{path_dir_txt}\cities.txt")
            with open(file_path) as f:
                lines = f.readlines()

            str = combo_city.get()
            pattern = re.compile(re.escape(str))
            with open(file_path, 'w') as f:
                for line in lines:
                    result = pattern.search(line)
                    if result is None:
                        f.write(line)

            upd_cities = r_cities_txt()
            combo_city.config(values=upd_cities)
            combo_city.current(2)

    def clear_for_fill():
        e_director.delete(0, END)
        e_doc.delete(0, END)
        e_directors_job.delete(0, END)
        e_directors_name.delete(0, END)
        e_mail.delete(0, END)

    def clear_data():
        e_fio.delete(0, END)
        e_director.delete(0, END)
        e_doc.delete(0, END)
        e_job.delete(0, END)
        e_name.delete(0, END)
        e_seria.delete(0, END)
        e_number.delete(0, END)
        e_point_of_get.delete(0, END)
        e_directors_job.delete(0, END)
        e_directors_name.delete(0, END)
        e_date.delete(0, END)

        e_code.delete(0, END)
        e_birth.delete(0, END)
        e_place_of_birth.delete(0, END)

        e_full_fio.delete(0, END)
        e_inn.delete(0, END)
        e_snils.delete(0, END)
        e_mail.delete(0, END)

    def select():
        nonlocal gender_data
        l = gender.get()
        if l == 1:
            gender_data = 'М'
        elif l == 2:
            gender_data = 'Ж'

    def close_ad(city):
        w_cities_txt(city)
        upd_cities = r_cities_txt()
        combo_city.config(values=upd_cities)
        combo_city.current(len(upd_cities) - 1)

    def r_cities_txt():
        if not ch_foo.check_path(var.cities_path):
            os.makedirs(var.cities_path)
        path_dir_txt = Path(var.cities_path)
        if not path_dir_txt.exists():
            os.mkdir(path_dir_txt)

        file_path = Path(rf"{path_dir_txt}\cities.txt")
        if not file_path.exists():
            my_file = open(rf"{path_dir_txt}\cities.txt", "w+")
            for city in cities:
                my_file.write(city + f'\n')
            my_file.close()

        data = []
        r_file = open(rf"{path_dir_txt}\cities.txt", "r").readlines()
        for line in r_file:
            data.append(line)
        return data

    def w_cities_txt(temp):
        path_dir_txt = Path(var.cities_path)
        if not path_dir_txt.exists():
            os.mkdir(path_dir_txt)

        file_path = Path(rf"{path_dir_txt}\cities.txt")
        if not file_path.exists():
            my_file = open(rf"{path_dir_txt}\cities.txt", "w+")
            for city in cities:
                my_file.write(city + f'\n')
            my_file.write(temp + f'\n')
            my_file.close()
        else:
            my_file = open(rf"{path_dir_txt}\cities.txt", "a+")
            my_file.write(temp + f'\n')
            my_file.close()

    def update_cities():
        def close_win_ad():
            ad.destroy()
            win.deiconify()

        def add():
            city = e_new_city.get()
            if city != '':
                close_ad(city)
                ad.destroy()
                win.deiconify()
            else:
                mes.error("Ошибка ввода", "Вы ввели некорректное значение!")

        ad = Toplevel()
        ad.geometry('400x100+100+50')
        ad.title('Введите новый город')
        ad.grab_set()
        win.withdraw()
        ad.protocol("WM_DELETE_WINDOW", close_win_ad)
        ad.resizable(False, False)
        ad.config(bg="#F1EEE9")

        f0 = Frame(ad, bg="#F1EEE9")
        f0.pack(fill=BOTH, padx=10, pady=5)

        e_new_city = Entry(f0)
        e_new_city.pack(fill=X, padx=10, ipady=2, expand=True)

        btn_add_city = ttk.Button(f0, text="Принять", command=add)
        btn_add_city.pack(side=BOTTOM, padx=5)

    def get_request_from_mail():
        get_requests_window.requests_win(win)
        win.withdraw()

    def open_out_window():
        output_window.out_xlsx(win)
        win.withdraw()

    def full_exit():
        service.full_exit(win)

    def get_table():
        if service.open_table():
            label13['foreground'] = 'green'
        else:
            label13['foreground'] = 'red'

    win = Toplevel()
    win.geometry('1500x450+100+50')
    win.title('Внесение данных в таблицу')
    win.grab_set()
    # root.withdraw()
    win.protocol("WM_DELETE_WINDOW", on_closing)
    win.resizable(True, True)
    win.minsize(1300, 450)
    win.maxsize(1800, 550)
    win.config(bg="#F1EEE9")
    win.columnconfigure(index=0, minsize=550, weight=550, pad=2)
    win.columnconfigure(index=1, minsize=600, weight=550, pad=2)
    win.columnconfigure(index=2, weight=50, pad=1)

    f0 = Frame(win, bg="#F1EEE9")
    f0.grid(row=0, column=0, sticky=W + E, padx=5)

    f1 = Frame(win, bg="#F1EEE9")
    f1.grid(row=1, column=0, sticky=W + E, padx=5)

    f2 = Frame(win, bg="#F1EEE9")
    f2.grid(row=2, column=0, sticky=W + E, padx=5)

    f3 = Frame(win, bg="#F1EEE9")
    f3.grid(row=3, column=0, sticky=W + E, padx=5)

    f3_1 = Frame(win, bg="#F1EEE9")
    f3_1.grid(row=6, column=0, sticky=W + E, padx=5)

    f4 = Frame(win, bg="#F1EEE9")
    f4.grid(row=7, column=0, sticky=W + E, padx=5)

    f5 = Frame(win, bg="#F1EEE9")
    f5.grid(row=8, column=0, sticky=W + E, padx=5)

    f6 = Frame(win, bg="#F1EEE9")
    f6.grid(row=9, column=0, sticky=W + E, padx=5)

    f7 = Frame(win, bg="#F1EEE9")
    f7.grid(row=10, column=0, sticky=W + E, padx=5)

    f8 = Frame(win, bg="#F1EEE9")
    f8.grid(row=1, column=1, sticky=W + E, padx=5)

    f9 = Frame(win, bg="#F1EEE9")
    f9.grid(row=2, column=1, sticky=W + E, padx=5)

    f10 = Frame(win, bg="#F1EEE9")
    f10.grid(row=3, column=1, sticky=W + E, padx=5)

    f11 = Frame(win, bg="#F1EEE9")
    f11.grid(row=4, column=1, sticky=W + E, padx=5)

    f12 = Frame(win, bg="#F1EEE9")
    f12.grid(row=5, column=1, sticky=W + E, padx=5)

    f13 = Frame(win, bg="#F1EEE9")
    f13.grid(row=6, column=1, sticky=W + E, padx=5)

    f14 = Frame(win, bg="#F1EEE9")
    f14.grid(row=7, column=1, sticky=W + E, padx=5)

    f15 = Frame(win, bg="#F1EEE9")
    f15.grid(row=8, column=1, sticky=W + E, padx=5)

    f16 = Frame(win, bg="#F1EEE9")
    f16.grid(row=9, column=1, sticky=W + E, padx=5)

    f17 = Frame(win, bg="#F1EEE9")
    f17.grid(row=10, column=1, sticky=W + E, padx=5)

    f18 = Frame(win, bg="#F1EEE9")
    f18.grid(row=11, column=1, sticky=W + E, padx=5)

    f19 = Frame(win, bg="#F1EEE9")
    f19.grid(row=12, column=1, sticky=W + E, padx=5)

    f20 = Frame(win, bg="#F1EEE9")
    f20.grid(row=11, column=0, sticky=W + E, padx=5)

    f21 = Frame(win, bg="#F1EEE9")
    f21.grid(row=12, column=0, sticky=W + E, padx=5)

    f22 = Frame(win, bg="#F1EEE9")
    f22.grid(row=0, column=1, sticky=W + E, padx=5)

    label_0 = Label(f0, width=25, text='Город:')
    label_0.pack(side=LEFT, pady=5)

    btn_add_city = ttk.Button(f0, text="Другой", command=update_cities)
    btn_add_city.pack(side=LEFT, padx=5)

    cities = ['Ачинск', 'Канск', 'Красноярск', 'Лесосибирск', 'Минусинск', 'Норильск']
    actual_cities = r_cities_txt()
    combo_city = ttk.Combobox(f0, values=actual_cities)
    combo_city.current(2)
    combo_city.pack(side=LEFT, fill=X, padx=10, pady=5)

    btn_del_city = ttk.Button(f0, text="<- Удалить", command=del_city)
    btn_del_city.pack(side=LEFT, padx=2, pady=5)

    label_1 = Label(f1, width=25, text='Фамилия И.О.:')
    label_1.pack(side=LEFT, pady=5)

    e_fio = Entry(f1)
    e_fio.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_2 = Label(f2, width=25, text='В лице руководителя:')
    label_2.pack(side=LEFT, pady=5)

    e_director = Entry(f2)
    e_director.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_3 = Label(f3, width=25, text='На основании:')
    label_3.pack(side=LEFT, pady=5)

    e_doc = Entry(f3)
    e_doc.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_3_1 = Label(f3_1, width=25, text='Уполномочивает', background='grey', foreground='yellow')
    label_3_1.pack(side=TOP, pady=5)

    label_4 = Label(f4, width=25, text='Должность:')
    label_4.pack(side=LEFT, pady=5)

    e_job = Entry(f4)
    e_job.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_5 = Label(f5, width=25, text='Сотрудника (полн.ФИО):')
    label_5.pack(side=LEFT, pady=5)

    e_name = Entry(f5)
    e_name.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_17 = Label(f6, width=25, text='ИНН:')
    label_17.pack(side=LEFT, pady=5)

    e_inn = Entry(f6)
    e_inn.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_18 = Label(f7, width=25, text='СНИЛС:')
    label_18.pack(side=LEFT, pady=5)

    e_snils = Entry(f7)
    e_snils.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_16 = Label(f8, width=25, text='ФИО (полное):')
    label_16.pack(side=LEFT, pady=5)

    e_full_fio = Entry(f8)
    e_full_fio.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_6 = Label(f9, width=25, text='Серия паспорта:')
    label_6.pack(side=LEFT, pady=5)

    e_seria = Entry(f9)
    e_seria.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_7 = Label(f10, width=25, text='Номер паспорта:')
    label_7.pack(side=LEFT, pady=5)

    e_number = Entry(f10)
    e_number.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_9 = Label(f11, width=25, text='Кем выдан:')
    label_9.pack(side=LEFT, pady=5)

    e_point_of_get = Entry(f11)
    e_point_of_get.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_8 = Label(f12, width=25, text='Дата выдачи паспорта:')
    label_8.pack(side=LEFT, pady=5)

    e_date = Entry(f12)
    e_date.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_12 = Label(f13, width=25, text='Код места выдачи:')
    label_12.pack(side=LEFT, pady=5)

    e_code = Entry(f13)
    e_code.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_15 = Label(f14, width=25, text='Пол:')
    label_15.pack(side=LEFT, pady=5)

    gender = IntVar()

    male_checkbutton = Radiobutton(f14, text="М", value=1, variable=gender, padx=10, command=select)
    male_checkbutton.pack(side=LEFT, pady=5)

    female_checkbutton = Radiobutton(f14, text="Ж", value=2, variable=gender, padx=10, command=select)
    female_checkbutton.pack(side=LEFT, pady=5)

    label_13 = Label(f15, width=25, text='Дата рождения:')
    label_13.pack(side=LEFT, pady=5)

    e_birth = Entry(f15)
    e_birth.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_14 = Label(f16, width=25, text='Место рождения:')
    label_14.pack(side=LEFT, pady=5)

    e_place_of_birth = Entry(f16)
    e_place_of_birth.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_10 = Label(f17, width=25, text='Должность руководителя:')
    label_10.pack(side=LEFT, pady=5)

    e_directors_job = Entry(f17)
    e_directors_job.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_11 = Label(f18, width=25, text='И.О.Фамилия руководителя:')
    label_11.pack(side=LEFT, pady=5)

    e_directors_name = Entry(f18)
    e_directors_name.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    label_19 = Label(f19, width=25, text='Почта:')
    label_19.pack(side=LEFT, pady=5)

    e_mail = Entry(f19)
    e_mail.pack(fill=X, padx=10, ipady=2, expand=True, pady=5)

    btn_confirm = ttk.Button(f20, text="Принять", command=confirm_data)
    btn_confirm.pack(side=RIGHT, pady=5, padx=10)

    btn_clear_e = ttk.Button(f20, text="Очистить", command=clear_data)
    btn_clear_e.pack(side=LEFT, padx=5, pady=5)

    btn_fill_e = ttk.Button(f20, text="Заполнить", command=insert_data)
    btn_fill_e.pack(side=LEFT, padx=5, pady=5)

    btn_fill_e = ttk.Button(f20, text="Данные", command=ch_foo.open_xlsx_sample)
    btn_fill_e.pack(side=LEFT, padx=17, pady=5)

    label12 = ttk.Label(f21, text="Статус загрузки файлов:")
    label12.pack(side=LEFT, padx=10, pady=5)

    label13 = ttk.Label(f21, text="Таблица с данными", foreground='red', justify=LEFT)
    label13.pack(side=RIGHT, padx=10, pady=5)

    if var.path_table != '':
        label13['foreground'] = 'green'
    else:
        label13['foreground'] = 'red'

    btn_full_exit = ttk.Button(f22, text="Выйти из программы", command=full_exit)
    btn_full_exit.pack(side=RIGHT, padx=17, pady=5)

    main_menu = Menu(win)
    win.config(menu=main_menu)

    # Выбрать таблицу
    file_menu = Menu(main_menu, tearoff=0)
    file_menu.add_command(label="Выбрать таблицу c данными", command=get_table)
    file_menu.add_command(label="Очистить таблицу", command=clear_xlsx)
    file_menu.add_command(label="Создать таблицу", command=create_xlsx)
    main_menu.add_cascade(label="Таблица", menu=file_menu, )

    # Вывод значений
    out_menu = Menu(main_menu, tearoff=0)
    out_menu.add_command(label="Вывести значения таблицы", command=open_out_window)
    out_menu.add_command(label="Запросы из почты", command=get_request_from_mail)
    main_menu.add_cascade(label="Данные", menu=out_menu, )

    insert_data()