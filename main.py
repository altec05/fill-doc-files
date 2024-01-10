import tkinter.scrolledtext
import docxtpl
from docxtpl import DocxTemplate
import openpyxl
from tkinter import *
from tkinter import messagebox
import os
from pathlib import Path
from tkinter import ttk

import input_window, output_window, service, get_requests_window
import variables as var
from service import backup_bd
import messages as mes


def open_custom_fill():
    input_window.open_custom_fill(root)
    root.withdraw()


def open_out_window():
    output_window.out_xlsx(root)
    root.withdraw()


def get_request_from_mail():
    get_requests_window.requests_win(root)
    root.withdraw()


def start_backup():
    result = messagebox.askyesno('Резервное копирование', 'Начать резервное копирование файлов на сервер?')
    print(result)
    if result:
        backup_bd('')


def error(title, text):
    messagebox.showerror(title=title, message=text)


def info(title, text):
    messagebox.showinfo(title=title, message=text)


def get_dict():
    start = e_start.get().upper()
    end = e_end.get().upper()
    test = []
    wb = openpyxl.load_workbook(f'{var.path_table}')
    sheet = wb['Доверенность']

    rows = 0

    try:
        for row in sheet[f'{start}':f'{end}']:
            rows += 1
            for cellObj in row:
                if cellObj.value == None or cellObj.value == '':
                    continue
                test.append(cellObj.value)

        info('Обработка входных данных', "Данные из таблицы успешно загружены!\n\nПереходим к "
                                         "заполнению шаблона данными.\n"
                                         f"Было обработано строк: {rows}")
        wb.close()
        return test
    except ValueError:
        error("Обработка входных данных", "Введите корректные координаты!")
        wb.close()
    except:
        error("Обработка входных данных", "Ошибка обработки данных!")
        wb.close()


def start(test):
    x = 0
    files = 0
    path = Path(var.doverennost_path)

    if not path.exists():
        os.mkdir(path)

    try:
        while x < len(test):
            doc = DocxTemplate(f'{var.path_sample}')
            context = {'director': test[x + 2], 'doc': test[x + 3], 'job': test[x + 4], 'name_of_subject': test[x + 5],
                       'seria': test[x + 6], 'num': test[x + 7], 'date': test[x + 8], 'passport_from_1': test[x + 9],
                       'job_of_director': test[x + 10], 'name_of_director': test[x + 11]}
            doc.render(context)
            path_dirs = Path(f'{path}/{test[x]}')
            if not path_dirs.exists():
                os.mkdir(path_dirs)
            doc.save(f'{path}/{test[x]}/Доверенность ' + test[x + 1] + '.docx')
            x += 12  # прыжок на следующую строку
            files += 1
        messagebox.showinfo(title="Обработка входных данных", message="Шаблон успешно обработан и заполнен.\n"
                                                                      f"Создано файлов: {files}.")
    except:
        error("Обработка входных данных", "Ошибка заполнения шаблона!")


def check_text():
    if e_start.get().isalnum() and e_end.get().isalnum():
        return True
    else:
        return False


def fill_files():
    if check_text():
        print(var.path_sample)
        print(var.path_table)
        if var.path_sample != '' and var.path_table != '':
            start(get_dict())
            # e_start.delete(0, END)
            # e_end.delete(0, END)
        else:
            error("Ошибка входных данных", "Выберите шаблон и таблицу с данными!")
    else:
        error("Ошибка входных данных", "Введите начальную и конечную точку таблицы!")


def open_dir():
    path = Path(var.doverennost_path)

    if path.exists():
        os.startfile(path)
    else:
        error("Открытие папки", "Каталог не существует!")


def get_sample():
    if service.open_sample_doc():
        label3['foreground'] = 'green'
    else:
        label3['foreground'] = 'red'


def get_table():
    if service.open_table():
        label4['foreground'] = 'green'

        # Прописываем последние координаты в поле
        wb = openpyxl.load_workbook(var.path_table)
        sheet = wb['Доверенность']
        e_end.delete(0, END)
        e_end.insert(0, sheet['M'][-1].coordinate)
        wb.close()

        if var.path_table == '':
            mes.error("Таблица с данными", "Выберите таблицу для обработки!")
        else:
            wb = openpyxl.load_workbook(var.path_table)
            sheet = wb["Доверенность"]

            start = 'B2'
            end = sheet['C'][-1].coordinate
            text_row = ''
            from_tab_list = []
            rows = 0

            try:
                for row in sheet[f'{start}':f'{end}']:
                    row_number = rows + 2
                    # from_tab_list.append(row_number)
                    cell_row_number = ''
                    rows += 1
                    for cellObj in row:
                        if cellObj.value == None or cellObj.value == '':
                            continue
                        from_tab_list.append(str(cellObj.value))
                        cell_row_number = cellObj.row
                    from_tab_list.append(cell_row_number)

                i = 0
                for elem in from_tab_list:
                    if i == 3:
                        i = 0
                    if i <= 1:
                        text_row += str(elem) + '\t-\t'
                    else:
                        text_row += 'M' + str(elem) + '\n'
                    i += 1

                print(text_row)
                data_text['state'] = 'normal'
                data_text.delete("1.0", END)
                data_text.insert("1.0", text_row)
                data_text['state'] = 'disabled'
            except:
                mes.error("Обработка входных данных", "Ошибка при заполнении поля вывода!")
                wb.close()
        wb.close()
    else:
        label4['foreground'] = 'red'

root = Tk()
root.title('Доверенность')
root.geometry('385x275+100+50')
root.resizable(False, False)

main_menu = Menu(root)
root.config(menu=main_menu)

file_menu = Menu(main_menu, tearoff=0)
file_menu.add_command(label="Выбрать таблицу c данными", command=get_table)
file_menu.add_command(label="Выбрать шаблон", command=get_sample)
main_menu.add_cascade(label="Файлы", menu=file_menu, )

# Ручной ввод
fill_menu = Menu(main_menu, tearoff=0)
fill_menu.add_command(label="Добавить", command=open_custom_fill)
fill_menu.add_command(label="Вывод", command=open_out_window)
fill_menu.add_command(label="Запросы из почты", command=get_request_from_mail)
main_menu.add_cascade(label="Данные", menu=fill_menu, )

# Ручной ввод данных
service_menu = Menu(main_menu, tearoff=0)
service_menu.add_command(label="Провести резервное копирование на сервер", command=start_backup)
main_menu.add_cascade(label="Сервис", menu=service_menu, )

f1 = Frame(root)
f1.pack(fill=X, padx=10, pady=5)

f2 = Frame(root)
f2.pack(fill=X, padx=10, pady=5)

f3 = Frame(root)
f3.pack(fill=X, padx=10, pady=5)

f4 = Frame(root)
f4.pack(fill=X, padx=10, pady=5)

f5 = Frame(root)
f5.pack(fill=X, padx=10, pady=5)

label1 = ttk.Label(f1, text="Начальная точка:", width=17)
label1.pack(side=LEFT, padx=10)

e_start = ttk.Entry(f1)
e_start.pack(side=RIGHT, fill=X, expand=True)

label2 = ttk.Label(f2, text="Конечная точка:", width=17)
label2.pack(side=LEFT, padx=10)

e_end = ttk.Entry(f2)
e_end.pack(side=RIGHT, fill=X, expand=True)

btn_show_table = ttk.Button(f3, text="Заполнить", command=fill_files)
btn_show_table.pack(side=RIGHT)

btn_show_dir = ttk.Button(f3, text="Результат", command=open_dir)
btn_show_dir.pack(side=LEFT, padx=10)

label3_1 = ttk.Label(f4, text="Статус загрузки файлов: ")
label3_1.pack(side=LEFT, padx=10)

label3 = ttk.Label(f4, text="Шаблон", foreground='red', justify=LEFT)
label3.pack(side=BOTTOM, padx=10)

if var.path_table != '':
    label3['foreground'] = 'green'
else:
    label3['foreground'] = 'red'

label4 = ttk.Label(f4, text="Таблица с данными", foreground='red', justify=LEFT)
label4.pack(side=BOTTOM, padx=10)

if var.path_table != '':
    label4['foreground'] = 'green'
else:
    label4['foreground'] = 'red'

e_start.delete(0, END)
e_start.insert(0, 'B2')

label5 = ttk.Label(f5, text="Содержимое таблицы с данными:", justify=LEFT)
label5.pack(side=TOP, padx=10)

data_text = tkinter.scrolledtext.ScrolledText(f5, width=50,  height=10, state=DISABLED)
data_text.pack(fill=BOTH, side=LEFT, expand=True)

root.mainloop()
